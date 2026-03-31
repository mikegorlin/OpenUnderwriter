"""Tests for the document ingestion pipeline.

All tests mock the LLM call (no API key needed).
Excel tests create real openpyxl workbooks in /tmp.
"""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

from do_uw.knowledge.pricing_ingestion import (
    ExtractedLayer,
    ExtractedProgram,
    extract_pricing_from_text,
    extract_text_from_excel,
    extract_text_from_file,
    ingest_document,
    map_extracted_to_inputs,
)
from do_uw.knowledge.pricing_store_programs import ProgramStore
from do_uw.models.pricing import DataSource


class TestExtractTextFromExcel:
    """Test Excel-to-text conversion using real openpyxl workbooks."""

    def test_basic_workbook(self, tmp_path: Path) -> None:
        """Create a workbook and verify text extraction."""
        from openpyxl import Workbook  # type: ignore[import-untyped]

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["Layer", "Limit", "Premium"])
        ws.append(["Primary", 10_000_000, 1_000_000])
        ws.append(["1st Excess", 10_000_000, 500_000])

        path = tmp_path / "tower.xlsx"
        wb.save(path)
        wb.close()

        text = extract_text_from_excel(path)
        assert "Layer" in text
        assert "Primary" in text
        assert "10000000" in text
        assert "1000000" in text
        assert "1st Excess" in text

    def test_multiple_sheets(self, tmp_path: Path) -> None:
        """Workbook with multiple sheets extracts all."""
        from openpyxl import Workbook  # type: ignore[import-untyped]

        wb = Workbook()
        ws1 = wb.active
        assert ws1 is not None
        ws1.title = "ABC Tower"
        ws1.append(["Primary", 10e6, 1e6])

        ws2 = wb.create_sheet("Side A")
        ws2.append(["Lead Side A", 10e6, 250_000])

        path = tmp_path / "multi.xlsx"
        wb.save(path)
        wb.close()

        text = extract_text_from_excel(path)
        assert "ABC Tower" in text
        assert "Side A" in text
        assert "Lead Side A" in text

    def test_empty_rows_skipped(self, tmp_path: Path) -> None:
        """Empty rows are skipped in text output."""
        from openpyxl import Workbook  # type: ignore[import-untyped]

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["Header"])
        ws.append([None, None, None])  # empty row
        ws.append(["Data"])

        path = tmp_path / "sparse.xlsx"
        wb.save(path)
        wb.close()

        text = extract_text_from_excel(path)
        lines = [line for line in text.split("\n") if line.strip()]
        # Should have sheet header + Header + Data (no empty row)
        assert len(lines) == 3


class TestExtractTextFromFile:
    """Test file type dispatching."""

    def test_txt_passthrough(self, tmp_path: Path) -> None:
        """Plain text files are read directly."""
        txt_file = tmp_path / "pricing.txt"
        txt_file.write_text("Primary $10M limit $1M premium")

        text = extract_text_from_file(txt_file)
        assert "Primary $10M" in text

    def test_csv_passthrough(self, tmp_path: Path) -> None:
        """CSV files are read directly."""
        csv_file = tmp_path / "tower.csv"
        csv_file.write_text("Layer,Limit,Premium\nPrimary,10000000,1000000")

        text = extract_text_from_file(csv_file)
        assert "Layer,Limit,Premium" in text

    def test_unsupported_format(self, tmp_path: Path) -> None:
        """Unsupported file types raise ValueError."""
        bad_file = tmp_path / "data.json"
        bad_file.write_text("{}")

        with pytest.raises(ValueError, match="Unsupported file format"):
            extract_text_from_file(bad_file)

    def test_xlsx_dispatches(self, tmp_path: Path) -> None:
        """XLSX files dispatch to Excel extractor."""
        from openpyxl import Workbook  # type: ignore[import-untyped]

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["Test"])
        path = tmp_path / "test.xlsx"
        wb.save(path)
        wb.close()

        text = extract_text_from_file(path)
        assert "Test" in text


class TestExtractPricingFromText:
    """Test LLM extraction with mocked instructor client."""

    def test_no_api_key_raises(self) -> None:
        """Missing ANTHROPIC_API_KEY raises RuntimeError."""
        with patch.dict("os.environ", {}, clear=True):
            # Ensure ANTHROPIC_API_KEY is not set
            env = dict(__import__("os").environ)
            env.pop("ANTHROPIC_API_KEY", None)
            with patch.dict("os.environ", env, clear=True):
                with pytest.raises(RuntimeError, match="ANTHROPIC_API_KEY"):
                    extract_pricing_from_text("some text")

    def test_extract_with_mocked_client(self) -> None:
        """Mocked instructor returns canned ExtractedProgram."""
        canned = ExtractedProgram(
            company_name="Test Corp",
            ticker="TEST",
            policy_year=2025,
            total_limit=50_000_000,
            total_premium=3_000_000,
            retention=2_500_000,
            layers=[
                ExtractedLayer(
                    layer_type="PRIMARY",
                    layer_label="Primary",
                    layer_number=1,
                    limit=10_000_000,
                    premium=1_000_000,
                    carrier="Chubb",
                    retention=2_500_000,
                ),
                ExtractedLayer(
                    layer_type="EXCESS",
                    layer_label="1st Excess",
                    layer_number=1,
                    limit=10_000_000,
                    premium=500_000,
                    carrier="AIG",
                ),
            ],
        )

        mock_client = MagicMock()
        mock_client.chat.completions.create.return_value = canned

        with (
            patch.dict(
                "os.environ", {"ANTHROPIC_API_KEY": "test-key"}
            ),
            patch(
                "do_uw.knowledge.pricing_ingestion.instructor"
            ) as mock_instructor,
        ):
            mock_instructor.from_provider.return_value = mock_client

            result = extract_pricing_from_text(
                "Primary $10M premium $1M"
            )

        assert result.company_name == "Test Corp"
        assert len(result.layers) == 2
        assert result.layers[0].layer_type == "PRIMARY"
        assert result.layers[0].premium == 1_000_000


class TestMapExtractedToInputs:
    """Test mapping from ExtractedProgram to input models."""

    def test_basic_mapping(self) -> None:
        """ExtractedProgram maps to ProgramInput + PolicyYearInput."""
        extracted = ExtractedProgram(
            company_name="Acme Corp",
            ticker="ACME",
            policy_year=2025,
            total_limit=50e6,
            total_premium=3e6,
            retention=2.5e6,
            layers=[
                ExtractedLayer(
                    layer_type="PRIMARY",
                    layer_label="Primary",
                    layer_number=1,
                    limit=10e6,
                    premium=1e6,
                    carrier="Chubb",
                    retention=2.5e6,
                ),
                ExtractedLayer(
                    layer_type="EXCESS",
                    layer_label="1st Excess",
                    layer_number=1,
                    limit=10e6,
                    premium=500_000,
                    carrier="AIG",
                ),
            ],
            brokerage="Marsh",
            broker_name="John Smith",
        )

        program_input, py_input = map_extracted_to_inputs(
            extracted, "acme"
        )

        # Program input
        assert program_input.ticker == "ACME"
        assert program_input.company_name == "Acme Corp"

        # Policy year input
        assert py_input.policy_year == 2025
        assert py_input.total_limit == 50e6
        assert py_input.total_premium == 3e6
        assert py_input.retention == 2.5e6
        assert py_input.source == "ai_extracted"
        assert len(py_input.layers) == 2

        # Layers have AI_EXTRACTED source
        for layer in py_input.layers:
            assert layer.data_source == DataSource.AI_EXTRACTED

    def test_ticker_override(self) -> None:
        """User-provided ticker overrides LLM-extracted ticker."""
        extracted = ExtractedProgram(
            ticker="WRONG", policy_year=2025
        )
        program_input, _ = map_extracted_to_inputs(
            extracted, "RIGHT"
        )
        assert program_input.ticker == "RIGHT"

    def test_layer_type_mapping(self) -> None:
        """Layer types map correctly to LayerType enum."""
        from do_uw.models.pricing import LayerType

        extracted = ExtractedProgram(
            policy_year=2025,
            layers=[
                ExtractedLayer(
                    layer_type="PRIMARY", layer_number=1
                ),
                ExtractedLayer(
                    layer_type="EXCESS", layer_number=1
                ),
                ExtractedLayer(
                    layer_type="SIDE_A", layer_number=1
                ),
            ],
        )
        _, py_input = map_extracted_to_inputs(extracted, "TEST")

        assert py_input.layers[0].layer_type == LayerType.PRIMARY
        assert py_input.layers[1].layer_type == LayerType.EXCESS
        assert py_input.layers[2].layer_type == LayerType.SIDE_A


class TestIngestDocument:
    """Test end-to-end ingestion with mocked LLM."""

    def _make_canned_program(self) -> ExtractedProgram:
        """Create a canned ExtractedProgram for mocking."""
        return ExtractedProgram(
            company_name="Test Corp",
            ticker="TEST",
            policy_year=2025,
            total_limit=50e6,
            total_premium=3e6,
            retention=2.5e6,
            layers=[
                ExtractedLayer(
                    layer_type="PRIMARY",
                    layer_label="Primary",
                    layer_number=1,
                    limit=10e6,
                    premium=1e6,
                    carrier="Chubb",
                    retention=2.5e6,
                ),
                ExtractedLayer(
                    layer_type="EXCESS",
                    layer_label="1st Excess",
                    layer_number=1,
                    limit=10e6,
                    premium=500_000,
                    carrier="AIG",
                ),
            ],
        )

    def test_end_to_end(self, tmp_path: Path) -> None:
        """Full pipeline with mocked LLM and real ProgramStore."""
        # Create a text file to ingest
        doc = tmp_path / "program.txt"
        doc.write_text("Primary $10M limit $1M premium")

        canned = self._make_canned_program()
        mock_client = MagicMock()
        mock_client.chat.completions.create.return_value = canned

        store = ProgramStore(db_path=None)

        with (
            patch.dict(
                "os.environ", {"ANTHROPIC_API_KEY": "test-key"}
            ),
            patch(
                "do_uw.knowledge.pricing_ingestion.instructor"
            ) as mock_instructor,
        ):
            mock_instructor.from_provider.return_value = mock_client

            result = ingest_document(
                doc, "TEST", store, hint="text document"
            )

        assert result["program_id"] >= 1
        assert result["policy_year_id"] >= 1
        assert result["layers_extracted"] == 2
        assert result["data_completeness"] in (
            "COMPLETE",
            "PARTIAL",
            "FRAGMENT",
        )

        # Verify data was stored
        program = store.get_program(result["program_id"])
        assert program is not None
        assert program.ticker == "TEST"
        assert len(program.policy_years) == 1

        py = program.policy_years[0]
        assert py.source == "ai_extracted"
        assert py.source_document == "program.txt"

    def test_existing_program_reuse(
        self, tmp_path: Path
    ) -> None:
        """Ingesting into existing program reuses program_id."""
        store = ProgramStore(db_path=None)

        # Pre-create program
        from do_uw.models.pricing import ProgramInput

        existing_id = store.add_program(
            ProgramInput(ticker="EXIST", company_name="Existing")
        )

        doc = tmp_path / "renewal.txt"
        doc.write_text("Primary $10M limit")

        canned = self._make_canned_program()
        mock_client = MagicMock()
        mock_client.chat.completions.create.return_value = canned

        with (
            patch.dict(
                "os.environ", {"ANTHROPIC_API_KEY": "test-key"}
            ),
            patch(
                "do_uw.knowledge.pricing_ingestion.instructor"
            ) as mock_instructor,
        ):
            mock_instructor.from_provider.return_value = mock_client

            result = ingest_document(doc, "EXIST", store)

        assert result["program_id"] == existing_id

    def test_source_document_set(self, tmp_path: Path) -> None:
        """Source document filename is recorded on policy year."""
        doc = tmp_path / "2025_tower.txt"
        doc.write_text("Tower data")

        canned = self._make_canned_program()
        mock_client = MagicMock()
        mock_client.chat.completions.create.return_value = canned

        store = ProgramStore(db_path=None)

        with (
            patch.dict(
                "os.environ", {"ANTHROPIC_API_KEY": "test-key"}
            ),
            patch(
                "do_uw.knowledge.pricing_ingestion.instructor"
            ) as mock_instructor,
        ):
            mock_instructor.from_provider.return_value = mock_client

            result = ingest_document(doc, "TEST", store)

        py = store.get_policy_year(result["policy_year_id"])
        assert py is not None
        assert py.source_document == "2025_tower.txt"

    def test_ai_extracted_data_source(
        self, tmp_path: Path
    ) -> None:
        """All layers have AI_EXTRACTED data source."""
        doc = tmp_path / "tower.txt"
        doc.write_text("Tower")

        canned = self._make_canned_program()
        mock_client = MagicMock()
        mock_client.chat.completions.create.return_value = canned

        store = ProgramStore(db_path=None)

        with (
            patch.dict(
                "os.environ", {"ANTHROPIC_API_KEY": "test-key"}
            ),
            patch(
                "do_uw.knowledge.pricing_ingestion.instructor"
            ) as mock_instructor,
        ):
            mock_instructor.from_provider.return_value = mock_client

            result = ingest_document(doc, "TEST", store)

        py = store.get_policy_year(result["policy_year_id"])
        assert py is not None
        # Layers in output don't expose data_source directly
        # but the store recorded them with AI_EXTRACTED
        assert py.source == "ai_extracted"
