"""Document ingestion pipeline for D&O pricing data.

Converts Excel tower spreadsheets, PDF program summaries, and text
files into structured pricing data via LLM extraction (instructor
library with DeepSeek). All LLM-extracted values are marked
with DataSource.AI_EXTRACTED for transparency (D7 decision).

Functions:
    extract_text_from_excel: Convert Excel to text for LLM.
    extract_text_from_pdf: Convert PDF to text via pdfplumber.
    extract_text_from_file: Dispatcher by file extension.
    extract_pricing_from_text: LLM extraction via instructor.
    ingest_document: End-to-end pipeline.
"""

from __future__ import annotations

import logging
import os
from pathlib import Path
from typing import Any, cast

from pydantic import BaseModel, Field

from do_uw.models.pricing import (
    DataCompleteness,
    DataSource,
    EnhancedLayerInput,
    LayerType,
    PolicyYearInput,
    ProgramInput,
    QuoteStatus,
)

try:
    import instructor as instructor
    from instructor import Mode
    import openai as openai
except ImportError:  # pragma: no cover
    instructor = None  # type: ignore[assignment]
    openai = None  # type: ignore[assignment]

logger = logging.getLogger(__name__)

# Default LLM model. Override via DO_UW_LLM_MODEL env var.
_DEFAULT_LLM_MODEL = os.environ.get("DO_UW_LLM_MODEL", "deepseek-chat")

# D&O tower system prompt for LLM extraction
_SYSTEM_PROMPT = """You are a D&O insurance pricing data extractor. Extract structured
tower pricing data from the provided document text.

D&O tower semantics:
- PRIMARY layer: Has a retention (SIR/deductible), a limit, and a premium. Only the
  primary has a retention. It is always the first layer.
- EXCESS layers: Stack above the primary. Named as 1st Excess, 2nd Excess, etc.
  Notation: "10xs10" means $10M excess of $10M (limit xs attachment).
  Each has a limit, premium, and carrier. No retention.
- SIDE_A layers: Sit ON TOP of the ABC tower, not parallel to it.
  Example: "10xs100 Lead Side A" = $10M Side A excess of $100M ABC tower.
  Side A = individual director/officer protection only (no entity coverage).

Common patterns:
- "10xs10" = $10M limit excess of $10M attachment
- Layer sizes are typically $5M or $10M increments
- Commission: Primary usually has commission; can be net 0%
- Rate = Premium / Limit
- Rate on Line (ROL) = Premium / Limit (same as rate)

Extract all layers, carriers, premiums, limits, and program-level data.
If a value is not clearly stated, leave it as null rather than guessing.
Dates should be in YYYY-MM-DD format.
Policy year should be the 4-digit year (e.g. 2025).
All monetary values should be in USD as raw numbers (no formatting)."""


class ExtractedLayer(BaseModel):
    """Layer extracted from a document by LLM."""

    layer_type: str = Field(description="PRIMARY, EXCESS, or SIDE_A")
    layer_label: str | None = Field(
        default=None,
        description="e.g. 'Primary', '1st Excess'",
    )
    layer_number: int = Field(description="Ordinal layer number")
    limit: float | None = Field(default=None, description="Layer limit in USD")
    premium: float | None = Field(default=None, description="Layer premium in USD")
    carrier: str | None = Field(default=None, description="Insurance carrier name")
    attachment: float | None = Field(default=None, description="Attachment point in USD")
    retention: float | None = Field(
        default=None,
        description="SIR/deductible (primary only)",
    )
    commission_pct: float | None = Field(default=None)
    am_best_rating: str | None = Field(default=None)


class ExtractedProgram(BaseModel):
    """Structured pricing data extracted from a document."""

    company_name: str | None = Field(default=None)
    ticker: str | None = Field(default=None)
    effective_date: str | None = Field(default=None, description="YYYY-MM-DD format")
    expiration_date: str | None = Field(default=None)
    policy_year: int | None = Field(default=None)
    total_limit: float | None = Field(default=None)
    total_premium: float | None = Field(default=None)
    retention: float | None = Field(default=None)
    layers: list[ExtractedLayer] = Field(default_factory=lambda: [])
    broker_name: str | None = Field(default=None)
    brokerage: str | None = Field(default=None)


def extract_text_from_excel(filepath: Path) -> str:
    """Convert Excel file to text for LLM extraction.

    Uses openpyxl to read .xlsx files cell by cell, converting
    each sheet to a tab-separated text representation.

    Args:
        filepath: Path to Excel file (.xlsx).

    Returns:
        Text representation of the workbook.

    Raises:
        ValueError: If file format is not supported by openpyxl.
    """
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as exc:
        msg = "openpyxl required for Excel parsing. Run: uv add openpyxl"
        raise RuntimeError(msg) from exc

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as exc:
        msg = (
            f"Cannot read {filepath.name}. "
            f"Only .xlsx format is supported by openpyxl. "
            f"For .xls files, convert to .xlsx first. Error: {exc}"
        )
        raise ValueError(msg) from exc

    text_parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        text_parts.append(f"=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                text_parts.append("\t".join(cells))

    wb.close()
    return "\n".join(text_parts)


def extract_text_from_pdf(filepath: Path) -> str:
    """Convert PDF file to text via pdfplumber.

    Extracts both text content and tables from each page,
    combining them into a single text representation.

    Args:
        filepath: Path to PDF file.

    Returns:
        Text representation of the PDF.
    """
    try:
        import pdfplumber
    except ImportError as exc:
        msg = "pdfplumber required for PDF parsing. Run: uv add pdfplumber"
        raise RuntimeError(msg) from exc

    text_parts: list[str] = []
    with pdfplumber.open(filepath) as pdf:
        for i, page in enumerate(pdf.pages):
            text_parts.append(f"=== Page {i + 1} ===")

            # Extract text content
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)

            # Extract tables as tab-separated
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    text_parts.append("--- Table ---")
                    for row in table:
                        cells = [str(c) if c is not None else "" for c in row]
                        text_parts.append("\t".join(cells))

    return "\n".join(text_parts)


_SUPPORTED_EXTENSIONS = {
    ".xlsx": "excel",
    ".xls": "excel",
    ".pdf": "pdf",
    ".txt": "text",
    ".md": "text",
    ".csv": "text",
}


def extract_text_from_file(filepath: Path) -> str:
    """Extract text from a file based on its extension.

    Dispatches to the appropriate extractor based on file type.

    Args:
        filepath: Path to the file.

    Returns:
        Text representation of the file content.

    Raises:
        ValueError: If file extension is not supported.
    """
    ext = filepath.suffix.lower()
    file_type = _SUPPORTED_EXTENSIONS.get(ext)

    if file_type is None:
        supported = ", ".join(sorted(_SUPPORTED_EXTENSIONS.keys()))
        msg = f"Unsupported file format '{ext}'. Supported formats: {supported}"
        raise ValueError(msg)

    if file_type == "excel":
        return extract_text_from_excel(filepath)
    if file_type == "pdf":
        return extract_text_from_pdf(filepath)

    # text, md, csv — read directly
    return filepath.read_text(encoding="utf-8")


def extract_pricing_from_text(text: str, hint: str | None = None) -> ExtractedProgram:
    """Extract structured pricing data from text via LLM.

    Uses the instructor library with DeepSeek to parse
    free-form document text into a structured ExtractedProgram.

    Args:
        text: Document text to extract from.
        hint: Optional context hint (e.g. "Excel tower spreadsheet").

    Returns:
        ExtractedProgram with structured pricing data.

    Raises:
        RuntimeError: If DEEPSEEK_API_KEY is not set.
    """
    api_key = os.environ.get("DEEPSEEK_API_KEY")
    if not api_key:
        msg = (
            "DEEPSEEK_API_KEY environment variable required "
            "for document ingestion. Set it with: "
            "export DEEPSEEK_API_KEY=your-key-here"
        )
        raise RuntimeError(msg)

    if instructor is None:
        msg = "instructor required for LLM extraction. Run: uv add instructor"
        raise RuntimeError(msg)

    if openai is None:
        msg = "openai required for DeepSeek API. Run: uv add openai"
        raise RuntimeError(msg)

    user_content = "Extract D&O insurance tower pricing from the following document:\n\n"
    if hint:
        user_content += f"Document type: {hint}\n\n"
    user_content += text

    raw_client = openai.OpenAI(
        api_key=api_key,
        base_url="https://api.deepseek.com",
        max_retries=10,
    )
    client = instructor.patch(raw_client, mode=Mode.TOOLS)

    result = cast(
        ExtractedProgram,
        client.chat.completions.create(
            model=_DEFAULT_LLM_MODEL,
            messages=[
                {"role": "system", "content": _SYSTEM_PROMPT},
                {"role": "user", "content": user_content},
            ],
            response_model=ExtractedProgram,
            max_tokens=4096,
            max_retries=2,
        ),
    )
    return result


def _map_layer_type(raw: str) -> LayerType:
    """Map extracted layer type string to LayerType enum."""
    upper = raw.upper().strip()
    if upper in ("PRIMARY", "PRI"):
        return LayerType.PRIMARY
    if upper in ("SIDE_A", "SIDEA", "SIDE A"):
        return LayerType.SIDE_A
    return LayerType.EXCESS


def _determine_completeness(
    layers: list[EnhancedLayerInput],
) -> DataCompleteness:
    """Determine data completeness from extracted layers."""
    if not layers:
        return DataCompleteness.FRAGMENT

    complete_count = 0
    for layer in layers:
        if layer.premium is not None and layer.limit_amount is not None:
            complete_count += 1

    if complete_count >= 2 and complete_count == len(layers):
        return DataCompleteness.COMPLETE
    if complete_count >= 2:
        return DataCompleteness.PARTIAL
    return DataCompleteness.FRAGMENT


def map_extracted_to_inputs(
    extracted: ExtractedProgram,
    ticker: str,
) -> tuple[ProgramInput, PolicyYearInput]:
    """Map ExtractedProgram to ProgramInput + PolicyYearInput.

    Args:
        extracted: LLM-extracted program data.
        ticker: User-provided ticker (authoritative override).

    Returns:
        Tuple of (ProgramInput, PolicyYearInput with layers).
    """
    # Build enhanced layers from extracted layers
    enhanced_layers: list[EnhancedLayerInput] = []
    for elayer in extracted.layers:
        layer_type = _map_layer_type(elayer.layer_type)
        enhanced_layers.append(
            EnhancedLayerInput(
                layer_type=layer_type,
                layer_label=elayer.layer_label,
                layer_number=elayer.layer_number,
                attachment_point=elayer.attachment,
                limit_amount=elayer.limit,
                premium=elayer.premium,
                carrier_name=elayer.carrier,
                carrier_rating=elayer.am_best_rating,
                commission_pct=elayer.commission_pct,
                data_source=DataSource.AI_EXTRACTED,
            )
        )

    completeness = _determine_completeness(enhanced_layers)

    # Build PolicyYearInput
    policy_year = extracted.policy_year or 2025
    py_input = PolicyYearInput(
        policy_year=policy_year,
        total_limit=extracted.total_limit,
        total_premium=extracted.total_premium,
        retention=extracted.retention,
        status=QuoteStatus.QUOTED,
        data_completeness=completeness,
        source="ai_extracted",
        layers=enhanced_layers,
    )

    # Build ProgramInput
    program_input = ProgramInput(
        ticker=ticker.upper(),
        company_name=extracted.company_name,
    )

    return program_input, py_input


def ingest_document(
    filepath: Path,
    ticker: str,
    store: Any,
    hint: str | None = None,
) -> dict[str, Any]:
    """End-to-end document ingestion pipeline.

    Extracts text from the file, uses LLM to parse structured
    pricing data, maps to Pydantic input models, and stores
    via ProgramStore.

    Args:
        filepath: Path to the document file.
        ticker: Stock ticker for the program (user authority).
        store: ProgramStore instance for persistence.
        hint: Optional document type hint for LLM context.

    Returns:
        Summary dict with program_id, policy_year_id,
        layers_extracted, and data_completeness.
    """
    # Step 1: Extract text
    text = extract_text_from_file(filepath)
    logger.info("Extracted %d chars from %s", len(text), filepath.name)

    # Step 2: LLM extraction
    extracted = extract_pricing_from_text(text, hint=hint)
    logger.info(
        "LLM extracted %d layers from %s",
        len(extracted.layers),
        filepath.name,
    )

    # Step 3: Map to input models
    program_input, py_input = map_extracted_to_inputs(extracted, ticker)

    # Set source document reference
    py_input.source_document = filepath.name

    # Step 4: Find or create program
    existing = store.get_program_by_ticker(ticker.upper())
    if existing is not None:
        program_id: int = existing.id
    else:
        program_id = store.add_program(program_input)

    # Step 5: Store policy year
    policy_year_id = store.add_policy_year(program_id, py_input)

    return {
        "program_id": program_id,
        "policy_year_id": policy_year_id,
        "layers_extracted": len(extracted.layers),
        "data_completeness": py_input.data_completeness.value,
    }
